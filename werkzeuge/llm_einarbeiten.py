"""Arbeitet KI-Beurteilungen in den Excel-Prüfbericht ein.

Aufruf:
  py werkzeuge\\llm_einarbeiten.py --bericht Pruefbericht_x.xlsx \\
     --beurteilungen llm_beurteilung.json

Erwartetes JSON (von der KI-Schicht erzeugt):
{
  "zusammenfassung": "3-6 Sätze Management Summary",
  "beurteilungen": [
    {"id": "K001", "urteil": "sachfremd-verdacht|privat-verdacht|"
                   "aktivierung-pruefen|doppelerfassung-verdacht|"
                   "unauffaellig|unklar",
     "begruendung": "1 Satz", "schwere": "hoch|mittel|hinweis|keine",
     "konfidenz": "hoch|mittel|niedrig"}
  ]
}
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

FONT = "Arial"
URTEIL_FILL = {
    "unauffaellig": PatternFill("solid", start_color="E2EFDA"),
    "unklar": PatternFill("solid", start_color="F2F2F2"),
}
VERDACHT_FILL = PatternFill("solid", start_color="FCE4D6")


def main(argv=None) -> int:
    p = argparse.ArgumentParser()
    p.add_argument("--bericht", required=True)
    p.add_argument("--beurteilungen", required=True)
    args = p.parse_args(argv)

    daten = json.loads(Path(args.beurteilungen).read_text(encoding="utf-8"))
    wb = load_workbook(args.bericht)

    ws = wb["KI-Kandidaten"]
    kopf = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    zeilen = {ws.cell(row=r, column=kopf["ID"]).value: r
              for r in range(2, ws.max_row + 1)}
    eingearbeitet = 0
    for b in daten.get("beurteilungen", []):
        zeile = zeilen.get(b.get("id"))
        if zeile is None:
            continue
        for feld, spalte in (("urteil", "KI-Urteil"),
                             ("begruendung", "KI-Begründung"),
                             ("schwere", "KI-Schwere"),
                             ("konfidenz", "KI-Konfidenz")):
            z = ws.cell(row=zeile, column=kopf[spalte], value=b.get(feld, ""))
            z.font = Font(name=FONT, size=10)
            z.alignment = Alignment(vertical="top", wrap_text=(feld == "begruendung"))
        urteil = (b.get("urteil") or "").strip().lower()
        zelle = ws.cell(row=zeile, column=kopf["KI-Urteil"])
        zelle.fill = URTEIL_FILL.get(urteil, VERDACHT_FILL)
        eingearbeitet += 1

    zusammenfassung = daten.get("zusammenfassung", "").strip()
    if zusammenfassung:
        ws_u = wb["Übersicht"]
        for r in range(1, ws_u.max_row + 1):
            if ws_u.cell(row=r, column=1).value == "KI-Zusammenfassung":
                anker = ws_u.cell(row=r + 1, column=1)
                anker.value = zusammenfassung
                anker.font = Font(name=FONT, size=10)
                anker.alignment = Alignment(vertical="top", wrap_text=True)
                break

    wb.save(args.bericht)
    print(f"{eingearbeitet} KI-Beurteilung(en) eingearbeitet"
          f"{' + Zusammenfassung' if zusammenfassung else ''}: {args.bericht}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

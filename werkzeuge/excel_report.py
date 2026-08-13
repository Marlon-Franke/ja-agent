"""Excel-Prüfbericht (openpyxl).

Bewusst ohne Formeln: Der Bericht ist ein statischer Befundbericht,
alle Werte sind deterministisch berechnet (keine Recalc-Abhängigkeit).
"""

from __future__ import annotations

import math
from collections import defaultdict
from datetime import date
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

import bilanz
from befunde import (EBENEN, KATALOG, KATALOG_BEREICHE, RANG, Befund,
                     Kontext, eur)

FONT = "Arial"
KOPF_FILL = PatternFill("solid", start_color="1F4E79")
REVIEW_FILL = PatternFill("solid", start_color="7030A0")
KOPF_FONT = Font(name=FONT, size=10, bold=True, color="FFFFFF")
STANDARD = Font(name=FONT, size=10)
LINK_FONT = Font(name=FONT, size=10, color="0563C1", underline="single")
SCHWERE_STIL = {
    "hoch": (PatternFill("solid", start_color="C00000"), Font(name=FONT, size=10, bold=True, color="FFFFFF")),
    "mittel": (PatternFill("solid", start_color="FFD966"), Font(name=FONT, size=10, color="000000")),
    "hinweis": (PatternFill("solid", start_color="BDD7EE"), Font(name=FONT, size=10, color="000000")),
}
BLATT_NAMEN = {
    "Datenintegrität": "Datenintegritaet",
    "Salden": "Salden",
    "AfA": "AfA",
    "USt/VSt": "USt-VSt",
    "Ausgangsrechnungen": "Rechnungen",
    "OPOS/Kreditoren": "OPOS-Kreditoren",
    "Bilanz (sonstige)": "Bilanz-Sonstige",
    "Vorjahresvergleich": "Vorjahr",
    "GuV-Plausibilität": "GuV-Plausis",
    "Ertragsteuer": "Ertragsteuer",
    "Cut-off": "Cut-off",
    "Personal/Privat": "Personal-Privat",
    "Gesellschafter": "Gesellschafter",
    "Statistik": "Statistik",
    "Fraud-Indikatoren": "Fraud-Indikatoren",
    "Stammdaten": "Stammdaten",
}

BEFUND_SPALTEN = [
    ("Schwere", 9), ("Check", 7), ("Prüfung", 28), ("Ebene", 13),
    ("Klasse", 7), ("Konto", 9), ("Kontobezeichnung", 20),
    ("Gegenkonto", 11), ("Datum", 11), ("Betrag (EUR)", 13), ("Beleg", 12),
    ("Buchungstext", 24), ("Befund", 55), ("Empfehlung", 42), ("Quelle", 17),
    ("KI", 4), ("Review-Status", 13), ("Bearbeiter", 12), ("Kommentar", 26),
]
KANDIDAT_SPALTEN = [
    ("ID", 6), ("Grund", 30), ("Konto", 9), ("Kontobezeichnung", 20),
    ("Gegenkonto", 11), ("Datum", 11), ("Betrag (EUR)", 13), ("BU", 5),
    ("Beleg", 13), ("Buchungstext", 32), ("Quelle", 17),
    ("KI-Urteil", 18), ("KI-Begründung", 52), ("KI-Schwere", 11),
    ("KI-Konfidenz", 12),
]


def _kopfzeile(ws, spalten, review_ab: int | None = None) -> None:
    for i, (name, breite) in enumerate(spalten, start=1):
        zelle = ws.cell(row=1, column=i, value=name)
        zelle.font = KOPF_FONT
        zelle.fill = REVIEW_FILL if review_ab and i >= review_ab else KOPF_FILL
        zelle.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = breite
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(spalten))}1"


def _zelle(ws, row, col, wert, wrap=False, fmt=None):
    z = ws.cell(row=row, column=col, value=wert)
    z.font = STANDARD
    z.alignment = Alignment(vertical="top", wrap_text=wrap)
    if fmt:
        z.number_format = fmt
    return z


def _befundblatt(wb: Workbook, titel: str, befunde: list[Befund], ctx: Kontext) -> None:
    ws = wb.create_sheet(titel[:31])
    _kopfzeile(ws, BEFUND_SPALTEN, review_ab=17)
    zeile = 2
    for f in sorted(befunde, key=lambda f: (RANG[f.schwere], f.check_id)):
        z = _zelle(ws, zeile, 1, f.schwere.upper())
        fill, font = SCHWERE_STIL[f.schwere]
        z.fill, z.font = fill, font
        z.alignment = Alignment(vertical="top", horizontal="center")
        _zelle(ws, zeile, 2, f.check_id)
        _zelle(ws, zeile, 3, f.check_name, wrap=True)
        _zelle(ws, zeile, 4, f"{f.ebene} {EBENEN[f.ebene]}")
        _zelle(ws, zeile, 5, f.klasse)
        _zelle(ws, zeile, 6, f.konto)
        _zelle(ws, zeile, 7, ctx.name(f.konto) if f.konto else "", wrap=True)
        _zelle(ws, zeile, 8, f.gegenkonto)
        _zelle(ws, zeile, 9, f.datum, fmt="DD.MM.YYYY")
        _zelle(ws, zeile, 10, float(f.betrag) if f.betrag is not None else None,
               fmt="#,##0.00")
        _zelle(ws, zeile, 11, f.beleg)
        _zelle(ws, zeile, 12, f.buchungstext, wrap=True)
        _zelle(ws, zeile, 13, f.text, wrap=True)
        _zelle(ws, zeile, 14, f.empfehlung, wrap=True)
        _zelle(ws, zeile, 15, f.quelle)
        _zelle(ws, zeile, 16, "ja" if f.llm_kandidat else "")
        for spalte in (17, 18, 19):
            _zelle(ws, zeile, spalte, "")
        hoehe = max(math.ceil(len(f.text) / 53), math.ceil(len(f.empfehlung) / 40), 1)
        if hoehe > 1:
            ws.row_dimensions[zeile].height = min(hoehe * 13 + 2, 92)
        zeile += 1


def _uebersicht(ws, ctx: Kontext, befunde: list[Befund], meta: dict) -> None:
    ws.title = "Übersicht"
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 46
    for spalte, breite in zip("CDEFGH", (13, 8, 8, 8, 9, 42)):
        ws.column_dimensions[spalte].width = breite
    titel = _zelle(ws, 1, 1, "Jahresabschluss-Prüfbericht (Entwurf – Arbeitshilfe)")
    titel.font = Font(name=FONT, size=15, bold=True, color="1F4E79")
    zeile = 3
    for label, wert in meta.items():
        _zelle(ws, zeile, 1, label).font = Font(name=FONT, size=10, bold=True)
        _zelle(ws, zeile, 2, wert, wrap=True)
        zeile += 1
    zeile += 1
    _zelle(ws, zeile, 1, "Kennzahlen").font = Font(name=FONT, size=12, bold=True)
    zeile += 1
    for label, wert in ctx.kennzahlen.items():
        _zelle(ws, zeile, 1, label).font = Font(name=FONT, size=10, bold=True)
        _zelle(ws, zeile, 2, wert)
        zeile += 1
    zeile += 1
    _zelle(ws, zeile, 1, "Prüfkatalog und Befundlage").font = Font(name=FONT, size=12, bold=True)
    zeile += 1
    _zelle(ws, zeile, 1,
           "Ebenen: 1 = technische Integrität, 2 = Regelprüfung, "
           "3 = Plausibilität, 4 = Anomalie · Klassen: R = regelbasiert, "
           "P = Plausibilität/Schwellenwert, A = Anomalie-Score, "
           "X = benötigt zusätzliche Datenquelle",
           wrap=True).font = Font(name=FONT, size=9, color="595959")
    ws.merge_cells(start_row=zeile, start_column=1, end_row=zeile, end_column=8)
    zeile += 1
    kopf = ("Check", "Prüfung", "Ebene", "Klasse", "hoch", "mittel", "hinweis", "Status")
    for i, name in enumerate(kopf, start=1):
        z = ws.cell(row=zeile, column=i, value=name)
        z.font, z.fill = KOPF_FONT, KOPF_FILL
    zeile += 1
    je_check: dict[str, dict[str, int]] = {}
    for f in befunde:
        je_check.setdefault(f.check_id, {"hoch": 0, "mittel": 0, "hinweis": 0})
        je_check[f.check_id][f.schwere] += 1
    for cid, name, _bereich, ebene, klasse in KATALOG:
        _zelle(ws, zeile, 1, cid)
        _zelle(ws, zeile, 2, name, wrap=True)
        _zelle(ws, zeile, 3, f"{ebene} {EBENEN[ebene]}")
        _zelle(ws, zeile, 4, klasse).alignment = Alignment(horizontal="center")
        zaehler = je_check.get(cid, {"hoch": 0, "mittel": 0, "hinweis": 0})
        for i, schwere in enumerate(("hoch", "mittel", "hinweis"), start=5):
            z = _zelle(ws, zeile, i, zaehler[schwere] or None)
            z.alignment = Alignment(horizontal="center")
            if zaehler[schwere]:
                z.fill, z.font = SCHWERE_STIL[schwere]
        if cid in ctx.geskippt:
            status = f"zusätzliche Prüfung – {ctx.geskippt[cid]}"
            _zelle(ws, zeile, 8, status, wrap=True)
        elif sum(zaehler.values()):
            blatt = BLATT_NAMEN.get(KATALOG_BEREICHE.get(cid, ""), "Alle Befunde")
            z = _zelle(ws, zeile, 8, f"Befunde – Blatt »{blatt}«", wrap=True)
            z.hyperlink = Hyperlink(ref=z.coordinate,
                                    location=f"'{blatt}'!A1")
            z.font = LINK_FONT
        else:
            _zelle(ws, zeile, 8, "geprüft, ohne Befund", wrap=True)
        zeile += 1
    zeile += 1
    _zelle(ws, zeile, 1, "KI-Zusammenfassung").font = Font(name=FONT, size=12, bold=True)
    zeile += 1
    ws.merge_cells(start_row=zeile, start_column=1, end_row=zeile + 4, end_column=8)
    platzhalter = _zelle(ws, zeile, 1,
        "– noch nicht erstellt; wird nach der KI-Durchsicht der Kandidaten "
        "ergänzt (werkzeuge/llm_einarbeiten.py) –", wrap=True)
    platzhalter.font = Font(name=FONT, size=10, italic=True, color="808080")
    zeile += 6
    ws.merge_cells(start_row=zeile, start_column=1, end_row=zeile + 3, end_column=8)
    _zelle(ws, zeile, 1,
        "Methodik: Deterministische Prüfung des DATEV-Buchungsstapels in vier "
        "Ebenen (Integrität, Regel, Plausibilität, Anomalie) – Regeln in Code, "
        "kein LLM; KI nur für die als 'KI' markierten Kandidaten. Stapeldaten "
        "enthalten Bruttoumsätze mit BU-Schlüsseln; GuV-Nettowerte und "
        "Steuerbeträge sind daraus rechnerisch abgeleitet. Kontengruppen gem. "
        "werkzeuge/konten_config.json (SKR-Standard, anpassbar). Prüfungen mit "
        "Status 'zusätzliche Prüfung' werden aktiv, sobald die genannte "
        "Datenquelle angeliefert wird. Diese Arbeitshilfe ersetzt keine "
        "fachliche Würdigung.",
        wrap=True).font = Font(name=FONT, size=9, color="595959")


_SCHWERE_FARBEN = (("hoch", "C00000"), ("mittel", "FFD966"),
                   ("hinweis", "BDD7EE"))


def _cockpit(wb: Workbook, ctx: Kontext, befunde: list[Befund]) -> None:
    """Visuelle Zusammenfassung: KPI-Block + gestapelte Balken je Bereich
    und Befunde je Prüfebene. Datentabellen bleiben sichtbar (Zahlenwerte
    sind nie nur über Farbe kodiert)."""
    ws = wb.create_sheet("Cockpit", index=1)
    ws.column_dimensions["A"].width = 24
    for spalte in "BCDE":
        ws.column_dimensions[spalte].width = 11
    titel = _zelle(ws, 1, 1, "Cockpit – Befundlage auf einen Blick")
    titel.font = Font(name=FONT, size=15, bold=True, color="1F4E79")

    zaehler = {s: sum(1 for f in befunde if f.schwere == s)
               for s, _farbe in _SCHWERE_FARBEN}
    mit_befund = {f.check_id for f in befunde}
    kpis = [
        ("Befunde gesamt", len(befunde), None),
        ("davon HOCH", zaehler["hoch"], "C00000"),
        ("davon MITTEL", zaehler["mittel"], "FFD966"),
        ("davon HINWEIS", zaehler["hinweis"], "BDD7EE"),
        ("KI-Kandidaten", sum(1 for f in befunde if f.llm_kandidat), None),
        ("Checks mit Befund", len(mit_befund), None),
        ("Checks ohne Befund",
         len(KATALOG) - len(mit_befund) - len(ctx.geskippt), None),
        ("zusätzliche Prüfungen (Datenquelle fehlt)", len(ctx.geskippt), None),
    ]
    zeile = 3
    for label, wert, farbe in kpis:
        _zelle(ws, zeile, 1, label).font = Font(name=FONT, size=10, bold=True)
        z = _zelle(ws, zeile, 2, wert)
        z.alignment = Alignment(horizontal="center")
        if farbe:
            z.fill = PatternFill("solid", start_color=farbe)
            if farbe == "C00000":
                z.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        zeile += 1

    # Datentabelle 1: Befunde je Bereich (Quelle des gestapelten Balkens)
    start1 = zeile + 2
    _zelle(ws, start1 - 1, 1, "Befunde je Bereich").font = Font(
        name=FONT, size=12, bold=True)
    for i, kopf in enumerate(("Bereich", "hoch", "mittel", "hinweis"), start=1):
        z = ws.cell(row=start1, column=i, value=kopf)
        z.font, z.fill = KOPF_FONT, KOPF_FILL
    bereiche = list(dict.fromkeys(b for _c, _n, b, _e, _k in KATALOG))
    r = start1 + 1
    for bereich in bereiche:
        _zelle(ws, r, 1, bereich)
        for i, (schwere, _farbe) in enumerate(_SCHWERE_FARBEN, start=2):
            anzahl = sum(1 for f in befunde
                         if f.bereich == bereich and f.schwere == schwere)
            _zelle(ws, r, i, anzahl).alignment = Alignment(horizontal="center")
        r += 1
    ende1 = r - 1

    chart1 = BarChart()
    chart1.type = "bar"
    chart1.grouping = "stacked"
    chart1.overlap = 100
    chart1.title = "Befunde je Bereich (nach Schweregrad)"
    chart1.add_data(Reference(ws, min_col=2, max_col=4, min_row=start1,
                              max_row=ende1), titles_from_data=True)
    chart1.set_categories(Reference(ws, min_col=1, min_row=start1 + 1,
                                    max_row=ende1))
    for serie, (_s, farbe) in zip(chart1.series, _SCHWERE_FARBEN):
        serie.graphicalProperties.solidFill = farbe
        serie.graphicalProperties.line.solidFill = "FFFFFF"
        serie.graphicalProperties.line.width = 12700  # 1 pt Trenner
    chart1.legend.position = "b"
    chart1.height = 12
    chart1.width = 22
    ws.add_chart(chart1, "G3")

    # Datentabelle 2: Befunde je Prüfebene
    start2 = ende1 + 3
    _zelle(ws, start2 - 1, 1, "Befunde je Prüfebene").font = Font(
        name=FONT, size=12, bold=True)
    for i, kopf in enumerate(("Ebene", "Befunde"), start=1):
        z = ws.cell(row=start2, column=i, value=kopf)
        z.font, z.fill = KOPF_FONT, KOPF_FILL
    r = start2 + 1
    for ebene in (1, 2, 3, 4):
        _zelle(ws, r, 1, f"{ebene} {EBENEN[ebene]}")
        anzahl = sum(1 for f in befunde if f.ebene == ebene)
        _zelle(ws, r, 2, anzahl).alignment = Alignment(horizontal="center")
        r += 1
    ende2 = r - 1

    chart2 = BarChart()
    chart2.type = "bar"
    chart2.title = "Befunde je Prüfebene"
    chart2.add_data(Reference(ws, min_col=2, min_row=start2, max_row=ende2),
                    titles_from_data=True)
    chart2.set_categories(Reference(ws, min_col=1, min_row=start2 + 1,
                                    max_row=ende2))
    chart2.series[0].graphicalProperties.solidFill = "1F4E79"
    chart2.legend = None
    chart2.height = 8
    chart2.width = 22
    ws.add_chart(chart2, "G28")


_ANLEITUNG = [
    ("Zweck", [
        f"Automatisch erzeugter Prüfbericht zum DATEV-Buchungsstapel: "
        f"{len(KATALOG)} deterministische Checks in vier Ebenen plus "
        "KI-Beurteilung "
        "ausgewählter Kandidaten. Arbeitshilfe – ersetzt keine fachliche "
        "Würdigung."]),
    ("Empfohlener Arbeitsablauf", [
        "1. Blatt »Cockpit«: Befundlage auf einen Blick (KPI und "
        "Diagramme je Bereich/Prüfebene); dann Blatt »Übersicht«: "
        "Befundmatrix – die Status-Spalte verlinkt direkt auf das "
        "jeweilige Detailblatt.",
        "2. Zuerst alle HOCH-Befunde klären (rot), danach Ebene 1 "
        "(technische Integrität) – sie trägt die Verlässlichkeit aller "
        "übrigen Ergebnisse.",
        "3. MITTEL-Befunde je Bereichsblatt abarbeiten, HINWEISE würdigen "
        "(AutoFilter in Zeile 1 nutzen).",
        "4. Blatt »KI-Kandidaten«: KI-Urteile mit Begründung und Konfidenz "
        "prüfen – die KI ergänzt, sie entscheidet nicht.",
        "5. Review-Spalten (lila Kopf: Review-Status, Bearbeiter, "
        "Kommentar) direkt in dieser Datei pflegen."]),
    ("Schweregrade", [
        "HOCH (rot) = sofort klären – klarer Regelverstoß oder starkes "
        "Fehlerindiz.",
        "MITTEL (gelb) = prüfen – erklärungsbedürftiger Sachverhalt.",
        "HINWEIS (blau) = würdigen – Auffälligkeit oder Risikosignal, "
        "kein Fehlernachweis."]),
    ("Ebenen und Klassen", [
        "Ebene 1 technische Integrität · 2 Regelprüfung · 3 Plausibilität "
        "· 4 Anomalie (nur Risikosignale).",
        "Klasse R = regelbasiert, P = Plausibilität/Schwellenwert, "
        "A = Anomalie-Score, X = benötigt zusätzliche Datenquelle.",
        "Status »geprüft, ohne Befund« = Check lief und fand nichts. "
        "Status »zusätzliche Prüfung« = wird aktiv, sobald die genannte "
        "Datenquelle bzw. Angabe geliefert wird (kein Mangel)."]),
    ("Spalten der Befundblätter", [
        "Quelle: bei buchungsgenauen Befunden Exportdatei:Zeilennummer "
        "(direktes Nachschlagen im Stapel); bei konten- oder "
        "abgleichsbasierten Befunden die Datenbasis des Checks "
        "(z. B. »OPOS-Liste ↔ Buchungsstapel«, »SuSa ↔ Buchungsstapel«, "
        "»Vorjahres-SuSa ↔ EB-Buchungen«) mit Kontonummer.",
        "KI = »ja«, wenn die Buchung zusätzlich der KI-Beurteilung "
        "vorgelegt wurde (Blatt »KI-Kandidaten«).",
        "Beträge in EUR; »netto« bezeichnet den rechnerisch um "
        "USt/VSt bereinigten Wert (Stapel führt Bruttoumsätze)."]),
]


def _anleitung(wb: Workbook, ctx: Kontext) -> None:
    ws = wb.create_sheet("Anleitung")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 110
    titel = _zelle(ws, 1, 1, "Anleitung zum Prüfbericht")
    titel.font = Font(name=FONT, size=15, bold=True, color="1F4E79")
    zeile = 3
    for abschnitt, zeilen in _ANLEITUNG:
        _zelle(ws, zeile, 1, abschnitt).font = Font(name=FONT, size=11, bold=True)
        for text in zeilen:
            z = _zelle(ws, zeile, 2, text, wrap=True)
            hoehe = max(1, math.ceil(len(text) / 105))
            if hoehe > 1:
                ws.row_dimensions[zeile].height = min(hoehe * 13 + 2, 60)
            zeile += 1
        zeile += 1
    _zelle(ws, zeile, 1, "Blätter dieses Berichts").font = Font(
        name=FONT, size=11, bold=True)
    zeile += 1
    for blatt in wb.sheetnames:
        if blatt == "Anleitung":
            continue
        z = _zelle(ws, zeile, 2, blatt)
        z.hyperlink = Hyperlink(ref=z.coordinate, location=f"'{blatt}'!A1")
        z.font = LINK_FONT
        zeile += 1
    wb.move_sheet("Anleitung", offset=-(len(wb.sheetnames) - 3))


def _kandidatenblatt(wb: Workbook, kandidaten: list[dict], ctx: Kontext) -> None:
    ws = wb.create_sheet("KI-Kandidaten")
    _kopfzeile(ws, KANDIDAT_SPALTEN, review_ab=12)
    for zeile, k in enumerate(kandidaten, start=2):
        _zelle(ws, zeile, 1, k["id"])
        _zelle(ws, zeile, 2, k["grund"], wrap=True)
        _zelle(ws, zeile, 3, k["konto"])
        _zelle(ws, zeile, 4, k["konto_name"], wrap=True)
        _zelle(ws, zeile, 5, k["gegenkonto"])
        datum = date.fromisoformat(k["datum"]) if k["datum"] else None
        _zelle(ws, zeile, 6, datum, fmt="DD.MM.YYYY")
        _zelle(ws, zeile, 7, k["betrag"], fmt="#,##0.00")
        _zelle(ws, zeile, 8, k["bu"])
        _zelle(ws, zeile, 9, k["beleg"])
        _zelle(ws, zeile, 10, k["buchungstext"], wrap=True)
        _zelle(ws, zeile, 11, k["quelle"])


_AGING_STUFEN = [(0, "nicht fällig"), (30, "0–30 Tage"), (60, "31–60 Tage"),
                 (90, "61–90 Tage"), (180, "91–180 Tage"),
                 (365, "181–365 Tage"), (99999, "> 365 Tage")]


def _aging_stufe(tage: int) -> str:
    if tage < 0:
        return "nicht fällig"
    for grenze, label in _AGING_STUFEN[1:]:
        if tage <= grenze:
            return label
    return "> 365 Tage"


def _opos_aging(wb: Workbook, ctx: Kontext) -> None:
    if not ctx.opos or ctx.datum_bis is None:
        return
    ws = wb.create_sheet("OPOS-Alterung")
    labels = [label for _g, label in _AGING_STUFEN]
    spalten = [("Konto", 10), ("Bezeichnung", 26), ("Art", 10)] + \
              [(label, 13) for label in labels] + [("Summe", 13), ("Posten", 8)]
    _kopfzeile(ws, spalten)
    je_konto: dict[int, dict[str, Decimal]] = defaultdict(lambda: defaultdict(Decimal))
    anzahl: dict[int, int] = defaultdict(int)
    for p in ctx.opos:
        datum = p.faellig or p.belegdatum
        stufe = _aging_stufe((ctx.datum_bis - datum).days) if datum else "0–30 Tage"
        je_konto[p.konto][stufe] += p.betrag
        anzahl[p.konto] += 1
    zeile = 2
    for konto in sorted(je_konto):
        art = "Debitor" if ctx.plan.ist_debitor(konto) else (
            "Kreditor" if ctx.plan.ist_kreditor(konto) else "Sachkonto")
        _zelle(ws, zeile, 1, konto)
        _zelle(ws, zeile, 2, ctx.name(konto), wrap=True)
        _zelle(ws, zeile, 3, art)
        gesamt = Decimal(0)
        for i, label in enumerate(labels, start=4):
            wert = je_konto[konto].get(label)
            if wert:
                gesamt += wert
            _zelle(ws, zeile, i, float(wert) if wert else None, fmt="#,##0.00")
        _zelle(ws, zeile, 4 + len(labels), float(gesamt), fmt="#,##0.00")
        _zelle(ws, zeile, 5 + len(labels), anzahl[konto])
        zeile += 1
    _zelle(ws, zeile + 1, 1,
           f"Alter = Stichtag ({ctx.datum_bis:%d.%m.%Y}) minus Fälligkeit "
           "(ersatzweise Belegdatum). Negative Beträge = Gutschriften.",
           wrap=True).font = Font(name=FONT, size=9, italic=True, color="595959")


def _saldenblatt(wb: Workbook, ctx: Kontext) -> None:
    ws = wb.create_sheet("Salden je Konto")
    spalten = [("Konto", 10), ("Bezeichnung", 30), ("Kontengruppen", 26),
               ("Soll (EUR)", 14), ("Haben (EUR)", 14), ("Saldo (EUR)", 14),
               ("Saldo netto (EUR)*", 15)]
    if ctx.susa_vj is not None:
        spalten.append(("Saldo Vorjahr (SuSa)", 16))
    spalten.append(("Buchungen", 10))
    spalten.append(("Bilanz-/GuV-Position", 38))
    _kopfzeile(ws, spalten)
    zeile = 2
    for k in sorted(ctx.anzahl):
        _zelle(ws, zeile, 1, k)
        _zelle(ws, zeile, 2, ctx.name(k), wrap=True)
        _zelle(ws, zeile, 3, ", ".join(ctx.plan.gruppen_von(k)), wrap=True)
        _zelle(ws, zeile, 4, float(ctx.soll[k]), fmt="#,##0.00")
        _zelle(ws, zeile, 5, float(ctx.haben[k]), fmt="#,##0.00")
        _zelle(ws, zeile, 6, float(ctx.saldo[k]), fmt="#,##0.00")
        _zelle(ws, zeile, 7, float(ctx.saldo_netto[k]), fmt="#,##0.00")
        spalte = 8
        if ctx.susa_vj is not None:
            vj = ctx.susa_vj.get(k)
            _zelle(ws, zeile, spalte, float(vj) if vj is not None else None,
                   fmt="#,##0.00")
            spalte += 1
        _zelle(ws, zeile, spalte, ctx.anzahl[k])
        _zelle(ws, zeile, spalte + 1, bilanz.position(ctx, k), wrap=True)
        zeile += 1
    # Implizite Steuerpositionen: die Netto-Bereinigung nimmt USt/VSt aus
    # den GuV-/AV-Salden heraus, ohne dass der Stapel Steuerkonten-Zeilen
    # enthält - diese beiden Zeilen schließen Bilanz und Saldensumme exakt.
    netto_spalte = 7
    pos_spalte = spalte + 1
    for label, wert, position in (
            ("Vorsteuer rechnerisch (implizit, nicht als Konto gebucht)",
             ctx.vst_rechnerisch, "E. Sonstige Aktiva"),
            ("Umsatzsteuer rechnerisch (implizit, nicht als Konto gebucht)",
             ctx.ust_rechnerisch, "C.IV. Sonstige Verbindlichkeiten")):
        if wert is None:
            continue
        vorzeichen = 1 if position.startswith("E.") else -1
        _zelle(ws, zeile, 1, "(implizit)")
        _zelle(ws, zeile, 2, label, wrap=True)
        _zelle(ws, zeile, netto_spalte, float(wert) * vorzeichen,
               fmt="#,##0.00")
        _zelle(ws, zeile, pos_spalte, position, wrap=True)
        zeile += 1
    _zelle(ws, zeile + 1, 1,
           "* Saldo netto: um rechnerisch abgeleitete USt/VSt bereinigt "
           "(nur GuV-/AV-Seite relevant); Bestandskonten brutto = netto. "
           "Die (implizit)-Zeilen weisen die herausgerechnete USt/VSt als "
           "Bilanzposition aus.",
           wrap=True).font = Font(name=FONT, size=9, italic=True, color="595959")


def _bilanz_guv(wb: Workbook, ctx: Kontext) -> None:
    """Bilanz- und GuV-Blatt, per SUMIF live mit »Salden je Konto«
    verknüpft (Positions-Spalte). Vorjahresspalte, wenn eine
    Vorjahres-SuSa übergeben wurde."""
    mit_vj = ctx.susa_vj is not None
    netto_sp = "G"
    vj_sp = "H" if mit_vj else None
    pos_sp = "J" if mit_vj else "I"
    quelle = "'Salden je Konto'"

    def summe(position: str, wert_sp: str, negiert: bool = False) -> str:
        formel = (f"SUMIF({quelle}!${pos_sp}:${pos_sp},"
                  f"\"{position}\",{quelle}!${wert_sp}:${wert_sp})")
        return f"=-{formel}" if negiert else f"={formel}"

    def blattkopf(ws, titel: str) -> None:
        ws.column_dimensions["A"].width = 52
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        z = _zelle(ws, 1, 1, titel)
        z.font = Font(name=FONT, size=15, bold=True, color="1F4E79")
        _zelle(ws, 2, 1,
               "Live verknüpft mit Blatt »Salden je Konto« "
               "(Spalte Bilanz-/GuV-Position); vereinfachte Gliederung nach "
               "§ 266/§ 275 HGB – amtliche Tiefe siehe Prüfkatalog Kap. 17."
               ).font = Font(name=FONT, size=9, italic=True, color="595959")
        for i, kopf in enumerate(
                ["Position", "Berichtsjahr (EUR)"] +
                (["Vorjahr (EUR)"] if mit_vj else []), start=1):
            zk = ws.cell(row=4, column=i, value=kopf)
            zk.font, zk.fill = KOPF_FONT, KOPF_FILL

    # --- GuV ---
    ws_guv = wb.create_sheet("GuV")
    blattkopf(ws_guv, "Gewinn- und Verlustrechnung (vereinfacht)")
    zeile = 5
    guv_start = zeile
    for pos in bilanz.GUV:
        negiert = pos in bilanz.GUV_ERTRAG
        _zelle(ws_guv, zeile, 1, pos)
        _zelle(ws_guv, zeile, 2, summe(pos, netto_sp, negiert),
               fmt="#,##0.00")
        if mit_vj:
            _zelle(ws_guv, zeile, 3, summe(pos, vj_sp, negiert),
                   fmt="#,##0.00")
        zeile += 1
    ergebnis_zeile = zeile + 1
    _zelle(ws_guv, ergebnis_zeile, 1,
           "Jahresüberschuss/Jahresfehlbetrag").font = Font(
        name=FONT, size=10, bold=True)
    for spalte in ("B",) + (("C",) if mit_vj else ()):
        z = _zelle(ws_guv, ergebnis_zeile, ord(spalte) - 64,
                   f"={spalte}{guv_start}-{spalte}{guv_start + 1}"
                   f"+{spalte}{guv_start + 2}"
                   f"-SUM({spalte}{guv_start + 3}:{spalte}{guv_start + 8})",
                   fmt="#,##0.00")
        z.font = Font(name=FONT, size=10, bold=True)

    # --- Bilanz ---
    ws_b = wb.create_sheet("Bilanz")
    blattkopf(ws_b, "Bilanz (vereinfacht)")
    zeile = 5
    _zelle(ws_b, zeile, 1, "AKTIVA").font = Font(name=FONT, size=11, bold=True)
    zeile += 1
    aktiv_start = zeile
    for pos in bilanz.AKTIV:
        _zelle(ws_b, zeile, 1, pos)
        _zelle(ws_b, zeile, 2, summe(pos, netto_sp), fmt="#,##0.00")
        if mit_vj:
            _zelle(ws_b, zeile, 3, summe(pos, vj_sp), fmt="#,##0.00")
        zeile += 1
    aktiv_summe = zeile
    _zelle(ws_b, aktiv_summe, 1, "Summe Aktiva").font = Font(
        name=FONT, size=10, bold=True)
    for spalte in ("B",) + (("C",) if mit_vj else ()):
        z = _zelle(ws_b, aktiv_summe, ord(spalte) - 64,
                   f"=SUM({spalte}{aktiv_start}:{spalte}{aktiv_summe - 1})",
                   fmt="#,##0.00")
        z.font = Font(name=FONT, size=10, bold=True)
    zeile = aktiv_summe + 2
    _zelle(ws_b, zeile, 1, "PASSIVA").font = Font(name=FONT, size=11, bold=True)
    zeile += 1
    passiv_start = zeile
    for pos in bilanz.PASSIV:
        _zelle(ws_b, zeile, 1, pos)
        _zelle(ws_b, zeile, 2, summe(pos, netto_sp, negiert=True),
               fmt="#,##0.00")
        if mit_vj:
            _zelle(ws_b, zeile, 3, summe(pos, vj_sp, negiert=True),
                   fmt="#,##0.00")
        zeile += 1
    _zelle(ws_b, zeile, 1, "Jahresergebnis lt. GuV (noch nicht verwendet)")
    _zelle(ws_b, zeile, 2, f"=GuV!B{ergebnis_zeile}", fmt="#,##0.00")
    if mit_vj:
        _zelle(ws_b, zeile, 3, f"=GuV!C{ergebnis_zeile}", fmt="#,##0.00")
    zeile += 1
    _zelle(ws_b, zeile, 1, "Saldenvortrags-Differenz (siehe SB-06)")
    _zelle(ws_b, zeile, 2, summe(bilanz.VORTRAG, netto_sp, negiert=True),
           fmt="#,##0.00")
    if mit_vj:
        _zelle(ws_b, zeile, 3, summe(bilanz.VORTRAG, vj_sp, negiert=True),
               fmt="#,##0.00")
    passiv_summe = zeile + 1
    _zelle(ws_b, passiv_summe, 1, "Summe Passiva").font = Font(
        name=FONT, size=10, bold=True)
    for spalte in ("B",) + (("C",) if mit_vj else ()):
        z = _zelle(ws_b, passiv_summe, ord(spalte) - 64,
                   f"=SUM({spalte}{passiv_start}:{spalte}{passiv_summe - 1})",
                   fmt="#,##0.00")
        z.font = Font(name=FONT, size=10, bold=True)
    kontrolle = passiv_summe + 2
    _zelle(ws_b, kontrolle, 1, "Kontrolle: Aktiva − Passiva (Soll: 0,00)")
    _zelle(ws_b, kontrolle, 2,
           f"=B{aktiv_summe}-B{passiv_summe}", fmt="#,##0.00")
    if mit_vj:
        _zelle(ws_b, kontrolle, 3,
               f"=C{aktiv_summe}-C{passiv_summe}", fmt="#,##0.00")


def schreibe_bericht(pfad, ctx: Kontext, befunde: list[Befund],
                     kandidaten: list[dict], meta: dict) -> None:
    wb = Workbook()
    _uebersicht(wb.active, ctx, befunde, meta)
    _cockpit(wb, ctx, befunde)
    _befundblatt(wb, "Alle Befunde", befunde, ctx)
    for bereich, blatt in BLATT_NAMEN.items():
        teil = [f for f in befunde if f.bereich == bereich]
        if teil:
            _befundblatt(wb, blatt, teil, ctx)
    _kandidatenblatt(wb, kandidaten, ctx)
    _opos_aging(wb, ctx)
    _bilanz_guv(wb, ctx)
    _saldenblatt(wb, ctx)
    _anleitung(wb, ctx)
    wb.save(pfad)
